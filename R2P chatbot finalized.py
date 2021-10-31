import streamlit as st
from chatterbot import ChatBot
from chatterbot.trainers import ListTrainer
from chatterbot.trainers import ChatterBotCorpusTrainer 
import openpyxl

#import json

def get_text():
    input_text = st.text_input("You: ")
    return input_text 

#data = json.loads(open(r'C:\Users\Jojo\Desktop\projects\chatbot\chatbot\chatbot\data_tolokers.json','r').read())#change path accordingly
#data2 = json.loads(open(r'C:\Users\Jojo\Desktop\projects\chatbot\chatbot\chatbot\sw.json','r').read())#change path accordingly

# tra = []
# for k, row in enumerate(data):
#     #print(k)
#     tra.append(row['dialog'][0]['text'])
# for k, row in enumerate(data2):
#     #print(k)
#     tra.append(row['dialog'][0]['text'])

wb = openpyxl.load_workbook("KDB.xlsx")
ws = wb.worksheets[0]

knowledgeDB = []
row_num = 1
for row in ws.iter_rows(min_row=1):
    if row[0].value is None:
        break
    question = ws["A" + str(row_num)]
    answer = ws["B" + str(row_num)]

    knowledgeDB.append(question.value)
    knowledgeDB.append(answer.value)
    row_num += 1

    

st.sidebar.title("R2P Kwodge Base Bot")
st.title("""
Bot is a knowledge sharing chatbot. Initialize the bot by clicking the "Initialize bot" button. 
""")


bot = ChatBot(name = 'chatbot', read_only = False,preprocessors=['chatterbot.preprocessors.clean_whitespace','chatterbot.preprocessors.convert_to_ascii','chatterbot.preprocessors.unescape_html'], logic_adapters = ['chatterbot.logic.BestMatch'])
#corpus_trainer = ChatterBotCorpusTrainer(bot) 
#corpus_trainer.train('chatterbot.corpus.english') 
#return bot

#bot = ChatBot("chatbot", read_only=False, 
            #   logic_adapters=[ 
            #                   {"import_path":"chatterbot.logic.BestMatch",
            #                    "default_response":"Sorry, I don't have an answer",
            #                    "maximum_similarity_threshold": 0.8
            #                    }
            #                   ])

list_to_train = knowledgeDB

# list_to_train = [
#                 "hi",
#                 "hi there",
#                 "what's your name?",
#                 "I'm a chatbot",
#                 "How are you?",
#                 "I am good",
#                 "What's your job?",
#                 "I'm here to answer your questions",
#                 "Where is T&E manual?",
#                 "Please refer to this site  https://alcon365.sharepoint.com/sites/InSight/SitePages/289/GPO/T%26E/TnE-JAPAN.aspx?cid=529ac87e-d98f-4cf8-afb8-2382fac22ece",
#                 "When is payment date?",
#                 "Please refer to this site  https://alcon365.sharepoint.com/:w:/r/sites/InSight/_layouts/15/Doc.aspx?sourcedoc=%7B128634DB-39C4-4E53-84F0-672C95CA4D19%7D&file=%E6%94%AF%E6%89%95%E6%97%A5%E7%A2%BA%E8%AA%8D%E6%96%B9%E6%B3%953.docx&action=default&mobileredirect=true",
#                 "支払い日はいつですか",
#                 "Please refer to this site   https://alcon365.sharepoint.com/:w:/r/sites/InSight/_layouts/15/Doc.aspx?sourcedoc=%7B128634DB-39C4-4E53-84F0-672C95CA4D19%7D&file=%E6%94%AF%E6%89%95%E6%97%A5%E7%A2%BA%E8%AA%8D%E6%96%B9%E6%B3%953.docx&action=default&mobileredirect=true",
#                  "ベンダー登録",
#                  "Please refer to Procurement sharepoint site.",
#     ]



ind = 1
if st.sidebar.button('Initialize bot'):
    #do something
    #bot = ChatBot(name = 'PyBot', read_only = False,preprocessors=['chatterbot.preprocessors.clean_whitespace','chatterbot.preprocessors.convert_to_ascii','chatterbot.preprocessors.unescape_html'], logic_adapters = ['chatterbot.logic.MathematicalEvaluation','chatterbot.logic.BestMatch'])
    # corpus_trainer = ChatterBotCorpusTrainer(bot) 
    # corpus_trainer.train('chatterbot.corpos.english') 
    # trainer2 = ListTrainer(bot) 
    # trainer2.train(tra)
    list_trainer = ListTrainer(bot)
    list_trainer.train(list_to_train)

    st.title("Your bot is ready to talk to you")
    ind = ind +1

user_input = get_text()


if True:
    st.text_area("Bot:", value=bot.get_response(user_input), height=200, max_chars=None, key=None)
else:
    st.text_area("Bot:", value="Please start the bot by clicking sidebar button", height=200, max_chars=None, key=None)
