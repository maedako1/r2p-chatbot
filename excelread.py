import openpyxl

wb = openpyxl.load_workbook("KDB.xlsx")
ws = wb["Sheet1"]


knowledgeDB = []
row_num = 1
for row in ws.iter_rows(min_row=1):
    if row[0].value is None:
        break
    question = ws["A" + str(row_num)]
    answer = ws["B" + str(row_num)]

    knowledgeDB.append(question.value)
    knowledgeDB.append(answer.value)
    row_num += 1

print(knowledgeDB)


